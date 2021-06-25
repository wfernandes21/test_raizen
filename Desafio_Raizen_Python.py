
# coding: utf-8

# ## INSTALAR 

# In[ ]:


## pip install pywin32
## pip instal pandas
## pip install numpy
## pip install python-dateutil
## pip install openpyxl


# ## Biblioteca

# In[ ]:


import win32com.client as win32
import os
import pandas as pd
import numpy as np
from pathlib import Path
import re
import sys
import urllib.request
from datetime import datetime
import dateutil.relativedelta
import time
import random
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
win32c = win32.constants


# ### Baixar arquivo da http://www.anp.gov.br

# In[ ]:


dir = './down_base'
try:
    os.makedirs(dir)
    print('diretorio criado')
except:
    print('diretorio já existe')


# In[ ]:


import urllib
urllib.request.urlretrieve("http://www.anp.gov.br/arquivos/dados-estatisticos/vendas-combustiveis/vendas-combustiveis-m3.xls", "./down_base/vendas-combustiveis-m3.xls")


# ### Abre e formata Excel

# In[ ]:


f_path = Path.cwd()
f_name = "./down_base/vendas-combustiveis-m3.xls"
filename = f_path / f_name
# create excel object
excel = win32.gencache.EnsureDispatch('Excel.Application')
# excel can be visible or not
excel.Visible = False  # False
wb = excel.Workbooks.Open(filename) 
excel.DisplayAlerts = False
wb.DoNotPromptForConvert = True
wb.CheckCompatibility = False
derivado = wb.Sheets("Plan1").PivotTables("Tabela dinâmica1")
disel = wb.Sheets("Plan1").PivotTables("Tabela dinâmica3")


# In[ ]:


derivado.ColumnGrand = True
derivado.RowGrand = True

disel.ColumnGrand = True
disel.RowGrand = True


# In[ ]:


wb.Sheets("Plan1").Range("X66").ShowDetail = True
wb.Sheets("Plan1").Range("k146").ShowDetail = True


# In[ ]:


DADOS_AT = [wb.Sheets("Plan1").Range("B42").Value]


# In[ ]:


### FECHA TUDO
wb.Close(True)
excel.Quit()


# ### Configura data da atualização

# In[ ]:


DADOS_AT = pd.DataFrame(DADOS_AT, columns =['at_data'])


# In[ ]:


DADOS_AT[['val1', 'val2', 'val3', 'val4', 'val5', 'val6', 'val7', 'val8']] = DADOS_AT['at_data'].str.split(' ', expand=True)
DADOS_AT['val8'] = DADOS_AT['val8'].str.replace(".","")


# In[ ]:


def mes(row):  
    mes = row['val6']

    if mes == "janeiro":
        return("1")
    elif mes == "fevereiro":
        return("2")
    elif mes == "março":
        return("3")
    elif mes == "abril":
        return("4")
    elif mes == "maio":
        return("5")
    elif mes == "junho":
        return("6")
    elif mes == "julho":
        return("7")
    elif mes == "agosto":
        return("8")
    elif mes == "setembro":
        return("9")
    elif mes == "outubro":
        return("10")
    elif mes == "novembro":
        return("11")
    elif mes == "dezembro":
        return("12")
    


# In[ ]:


DADOS_AT['val6'] = DADOS_AT.apply(mes, axis=1)


# In[ ]:


DADOS_AT['data_f'] = DADOS_AT['val4'] + "/" + DADOS_AT['val6'] + "/" + DADOS_AT['val8'] + " 00:00:00"


# In[ ]:


DADOS_AT['data_f'] = pd.to_datetime(DADOS_AT['data_f'])


# ### Baixando abas criadas no excel (automático)

# In[ ]:


try:
    DF_DERIVADO =pd.read_excel(open('./down_base/vendas-combustiveis-m3.xls', 'rb'),
                  sheet_name='Planilha1');  
    DF_DERIVADO = DF_DERIVADO.drop(columns=['TOTAL','REGIÃO']);
except:
    DF_DERIVADO =pd.read_excel(open('./down_base/vendas-combustiveis-m3.xls', 'rb'),
                  sheet_name='Plan2');  
    DF_DERIVADO = DF_DERIVADO.drop(columns=['TOTAL','REGIÃO']);


# In[ ]:


try:
    DF_DISEL =pd.read_excel(open('./down_base/vendas-combustiveis-m3.xls', 'rb'),
                  sheet_name='Planilha2');  
    DF_DISEL = DF_DISEL.drop(columns=['TOTAL','REGIÃO']);
except:
    DF_DISEL =pd.read_excel(open('./down_base/vendas-combustiveis-m3.xls', 'rb'),
                  sheet_name='Plan3');  
    DF_DISEL = DF_DISEL.drop(columns=['TOTAL','REGIÃO']);


# ### Modelando dados do DF_DISEL_FINAL

# In[ ]:


DF_DISEL_FINAL = DF_DISEL.melt(id_vars=["COMBUSTÍVEL", "ANO", "ESTADO", "UNIDADE" ], 
        var_name="month", 
        value_name="volume")


# In[ ]:


def mes2(row):  
    mes = row['month']

    if mes == "Jan":
        return("01")
    elif mes == "Fev":
        return("02")
    elif mes == "Mar":
        return("03")
    elif mes == "Abr":
        return("04")
    elif mes == "Mai":
        return("05")
    elif mes == "Jun":
        return("06")
    elif mes == "Jul":
        return("07")
    elif mes == "Ago":
        return("08")
    elif mes == "Set":
        return("09")
    elif mes == "Out":
        return("10")
    elif mes == "Nov":
        return("11")
    elif mes == "Dez":
        return("12")


# In[ ]:


DF_DISEL_FINAL['month'] = DF_DISEL_FINAL.apply(mes2, axis=1)


# In[ ]:


DF_DISEL_FINAL['year_month'] = DF_DISEL_FINAL['ANO'].astype(str) +"/"+ DF_DISEL_FINAL['month']
DF_DISEL_FINAL = DF_DISEL_FINAL.drop(columns=['ANO','month']);


# In[ ]:


day_t = DADOS_AT['data_f'].iloc[0]


# In[ ]:


DF_DISEL_FINAL['created_at'] = day_t


# In[ ]:


DF_DISEL_FINAL = DF_DISEL_FINAL.rename(columns={"ESTADO": "uf", "UNIDADE": "unit", "UNIDADE": "unit", "COMBUSTÍVEL": "product"})


# In[ ]:


DF_DISEL_FINAL = DF_DISEL_FINAL[[

'year_month',
'uf',
'product',
'unit',
'volume',
'created_at'

]]


# In[ ]:


DF_DISEL_FINAL = DF_DISEL_FINAL.sort_values(by=['year_month','uf','product'], ascending=True)


# In[ ]:


DF_DISEL_FINAL = DF_DISEL_FINAL.reset_index(drop=True)


# ### Modelando dados do DF_DERIVADO_FINAL

# In[ ]:


DF_DERIVADO_FINAL = DF_DERIVADO.melt(id_vars=["COMBUSTÍVEL", "ANO", "ESTADO", "UNIDADE" ], 
        var_name="month", 
        value_name="volume")


# In[ ]:


DF_DERIVADO_FINAL['month'] = DF_DERIVADO_FINAL.apply(mes2, axis=1)


# In[ ]:


DF_DERIVADO_FINAL['year_month'] = DF_DERIVADO_FINAL['ANO'].astype(str) +"/"+ DF_DERIVADO_FINAL['month']
DF_DERIVADO_FINAL = DF_DERIVADO_FINAL.drop(columns=['ANO','month']);


# In[ ]:


DF_DERIVADO_FINAL['created_at'] = day_t


# In[ ]:


DF_DERIVADO_FINAL = DF_DERIVADO_FINAL.rename(columns={"ESTADO": "uf", "UNIDADE": "unit", "UNIDADE": "unit", "COMBUSTÍVEL": "product"})


# In[ ]:


DF_DERIVADO_FINAL = DF_DERIVADO_FINAL[[

'year_month',
'uf',
'product',
'unit',
'volume',
'created_at'

]]


# In[ ]:


DF_DERIVADO_FINAL = DF_DERIVADO_FINAL.sort_values(by=['year_month','uf','product'], ascending=True)


# In[ ]:


DF_DERIVADO_FINAL = DF_DERIVADO_FINAL.reset_index(drop=True)


# ### Confirmando resultado agrupado

# In[ ]:


pd.pivot_table(DF_DISEL_FINAL,
   index=['year_month'],
   values=['volume'],
   columns=[ 'product'],
   fill_value='',
   aggfunc=np.sum,
   margins=True,
   margins_name='Total')


# In[ ]:


pd.pivot_table(DF_DERIVADO_FINAL,
   index=['year_month'],
   values=['volume'],
   columns=[ 'product'],
   fill_value='',
   aggfunc=np.sum,
   margins=True,
   margins_name='Total')


# ### Modelando dados do DF_BIG_TABLE

# In[ ]:


DF_BIG_1 = DF_DERIVADO_FINAL.copy()
DF_BIG_2 = DF_DISEL_FINAL.copy()


# In[ ]:


DF_BIG_1 ['DATA'] = 'DADOS_DERIVADO' 
DF_BIG_2 ['DATA'] = 'DADOS_DISEL'


# In[ ]:


DF_BIG_FINAL = DF_BIG_1.append(DF_BIG_2, ignore_index=True)


# ### Salvando no Excel

# In[ ]:


dir = './new_tables'
try:
    os.makedirs(dir)
    print('diretorio criado')
except:
    print('diretorio já existe')


# In[ ]:


# criar workbook
wb = Workbook('./new_tables/big_table.xlsx')
wb.save('./new_tables/big_table.xlsx')
wb = load_workbook('./new_tables/big_table.xlsx')


# In[ ]:


# adicionar workbook
sheet1 = wb.create_sheet('DF_DERIVADO_FINAL',0)
sheet2 = wb.create_sheet('DF_DISEL_FINAL',1)
sheet3 = wb.create_sheet('DF_BIG_FINAL',2)


# In[ ]:


# Remove existing sheet
ref = wb['Sheet']
wb.remove(ref)


# In[ ]:


# Seleciona sheet e exporta os dados

active = wb['DF_DERIVADO_FINAL']
for x in dataframe_to_rows(DF_DERIVADO_FINAL, index = False):
     active.append(x)
        
active = wb['DF_DISEL_FINAL']
for x in dataframe_to_rows(DF_DISEL_FINAL, index = False):
     active.append(x)

active = wb['DF_BIG_FINAL']
for x in dataframe_to_rows(DF_BIG_FINAL, index = False):
     active.append(x)


# In[ ]:


# Salva a excel
wb.save('./new_tables/big_table.xlsx')

