import os
import pandas as pd
import numpy as np
from pathlib import Path
import re
import sys
import urllib.request
from datetime import datetime
import dateutil.relativedelta
import time
from time import sleep
import random
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.model_selection import ParameterGrid
from requests_html import HTMLSession
import urllib
from airflow import DAG
from airflow.operators.bash_operator import BashOperator
from airflow.operators.python import PythonOperator
from airflow.models.xcom import XCom
import json


def baixa_dados_externo():
    ### Baixar arquivo da https://www.gov.br/anp/pt-br/centrais-de-conteudo

    dir = './dados_desafio_airflow/down_base'
    try:
        os.makedirs(dir)
        print('diretorio criado')
    except:
        print('diretorio já existe')

    data_link = pd.DataFrame({'link': ['https://dados.gov.br/dataset/vendas-de-derivados-de-petroleo-e-biocombustiveis/resource/c3d6e0b4-f86e-48f8-9325-6cc0d434b33f?inner_span=True', 'https://dados.gov.br/dataset/vendas-de-derivados-de-petroleo-e-biocombustiveis/resource/2429fdeb-df86-4e63-b248-2038f6c3e3cc?inner_span=True'],'tipo':['Derivados','Disel']})

    ## Ele acessa o site para pegar o link da ultima atualização
    df_links = pd.DataFrame()

    try:
        for i, row in data_link.iterrows():
            tipo = row['tipo']
            with HTMLSession() as s:
                df = pd.DataFrame()
                r = s.get(row['link'])
                hiddens = r.html.find('a', containing='https://www.gov.br/anp/pt-br/centrais-de-conteudo', first=True)#r.html.find('div[role=main]', first=True)
                df = pd.DataFrame([hiddens.links])
                df['tipo']= tipo
                df_links = df_links.append(df)
                print("dados {} deu bom !".format(row['link']))

    except:
        print("dados {} deu Ruim !".format(row['link']))

    # baixa csv direto do site
    try:
        for i, row in df_links.iterrows():
            link = row[0]
            tipo = row['tipo']
            urllib.request.urlretrieve(link, "./dados_desafio_airflow/down_base/dados_"+ tipo+".csv")
            print("Dados " + tipo + " Baixado")  
    except:
            print("Dados " + tipo + " Deu Ruim")  

    # Baixa arquivo pivotado apenas para saber a data da verção do arquivo
    urllib.request.urlretrieve("https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-estatisticos/de/vdpb/vendas-combustiveis-m3.xls/@@download/file/vendas-combustiveis-m3.xlsx", "./dados_desafio_airflow/down_base/vendas-combustiveis-m3.xlsx")
    print("Dados Pivotado Baixado") 

    book = openpyxl.load_workbook('./dados_desafio_airflow/down_base/vendas-combustiveis-m3.xlsx')
    sheet = book['Plan1']
    DADOS_AT = [sheet.cell(row=42, column=2).value]
    return DADOS_AT

def confirmando_data_atualizacao(ti):
    ### Configura data da atualização
    DADOS_AT = ti.xcom_pull(task_ids='baixa_dados_externo')
    DADOS_AT = pd.DataFrame(DADOS_AT, columns =['at_data'])

    DADOS_AT[['val1', 'val2', 'val3', 'val4', 'val5', 'val6', 'val7', 'val8']] = DADOS_AT['at_data'].str.split(' ', expand=True)
    DADOS_AT['val8'] = DADOS_AT['val8'].str.replace(".","")

    def mes(row):  
        mes = row['val6']

        if mes == "janeiro":
            return("1")
        elif mes == "fevereiro":
            return("2")
        elif mes == "março":
            return("3")
        elif mes == "abril":
            return("4")
        elif mes == "maio":
            return("5")
        elif mes == "junho":
            return("6")
        elif mes == "julho":
            return("7")
        elif mes == "agosto":
            return("8")
        elif mes == "setembro":
            return("9")
        elif mes == "outubro":
            return("10")
        elif mes == "novembro":
            return("11")
        elif mes == "dezembro":
            return("12")       

    DADOS_AT['val6'] = DADOS_AT.apply(mes, axis=1)
    list_data_f= [DADOS_AT['val4'] + "/" + DADOS_AT['val6'] + "/" + DADOS_AT['val8'] + " 00:00:00"][0][0]
    list_data_f = ti.xcom_push('list_data_key', list_data_f)

def tratando_dados_save_bigdate(ti):
    dfx = ti.xcom_pull(task_ids='confirmando_data_atualizacao', key='list_data_key')
    print(dfx)
    DADOS_AT = pd.DataFrame({'data_f':[dfx]})
    print(DADOS_AT)

    ### tratando dados

    DF_DERIVADO =pd.read_csv('./dados_desafio_airflow/down_base/dados_Derivados.csv', sep = ';');  
    DF_DERIVADO_FINAL = DF_DERIVADO.rename(columns={"UNIDADE DA FEDERAÇÃO": "uf", "PRODUTO": "product", 'MÊS':'month', 'ANO':'year', 'VENDAS': 'volume'})
    DF_DERIVADO_FINAL ['unit'] = 'm3'
    DF_DERIVADO_FINAL['volume'] = DF_DERIVADO_FINAL.volume.str.replace(',', '.').apply(lambda x: float(x))
    DF_DERIVADO_FINAL = DF_DERIVADO_FINAL[["product", "year", "uf", "unit", 'month','volume' ]]
    DF_DISEL =pd.read_csv('./dados_desafio_airflow/down_base/dados_Disel.csv', sep = ';');
    DF_DISEL_FINAL = DF_DISEL.rename(columns={"UNIDADE DA FEDERAÇÃO": "uf", "PRODUTO": "product", 'MÊS':'month', 'ANO':'year', 'VENDAS': 'volume'})
    DF_DISEL_FINAL ['unit'] = 'm3'
    DF_DISEL_FINAL['volume'] = DF_DISEL_FINAL.volume.str.replace(',', '.').apply(lambda x: float(x))
    DF_DISEL_FINAL = DF_DISEL_FINAL[["product", "year", "uf", "unit", 'month','volume' ]]
    ### Modelando dados do DF_DISEL_FINAL

    def mes2(row):  
        mes = row['month']

        if mes == "JAN":
            return("01")
        elif mes == "FEV":
            return("02")
        elif mes == "MAR":
            return("03")
        elif mes == "ABR":
            return("04")
        elif mes == "MAI":
            return("05")
        elif mes == "JUN":
            return("06")
        elif mes == "JUL":
            return("07")
        elif mes == "AGO":
            return("08")
        elif mes == "SET":
            return("09")
        elif mes == "OUT":
            return("10")
        elif mes == "NOV":
            return("11")
        elif mes == "DEZ":
            return("12")
    DF_DISEL_FINAL['month'] = DF_DISEL_FINAL.apply(mes2, axis=1)
    DF_DISEL_FINAL['year_month'] = DF_DISEL_FINAL['year'].astype(str) +"/"+ DF_DISEL_FINAL['month']
    DF_DISEL_FINAL = DF_DISEL_FINAL.drop(columns=['year','month']);
    day_t = DADOS_AT['data_f'].iloc[0]
    DF_DISEL_FINAL['created_at'] = day_t
    DF_DISEL_FINAL = DF_DISEL_FINAL.rename(columns={"ESTADO": "uf", "UNIDADE": "unit", "UNIDADE": "unit", "COMBUSTÍVEL": "product"})
    DF_DISEL_FINAL = DF_DISEL_FINAL[[
    'year_month',
    'uf',
    'product',
    'unit',
    'volume',
    'created_at'
    ]]
    DF_DISEL_FINAL = DF_DISEL_FINAL.sort_values(by=['year_month','uf','product'], ascending=True)
    DF_DISEL_FINAL = DF_DISEL_FINAL.reset_index(drop=True)
    ### Modelando dados do DF_DERIVADO_FINAL
    DF_DERIVADO_FINAL['month'] = DF_DERIVADO_FINAL.apply(mes2, axis=1)
    DF_DERIVADO_FINAL['year_month'] = DF_DERIVADO_FINAL['year'].astype(str) +"/"+ DF_DERIVADO_FINAL['month']
    DF_DERIVADO_FINAL = DF_DERIVADO_FINAL.drop(columns=['year','month']);
    DF_DERIVADO_FINAL['created_at'] = day_t
    DF_DERIVADO_FINAL = DF_DERIVADO_FINAL[[

    'year_month',
    'uf',
    'product',
    'unit',
    'volume',
    'created_at'

    ]]
    DF_DERIVADO_FINAL = DF_DERIVADO_FINAL.sort_values(by=['year_month','uf','product'], ascending=True)

    DF_DERIVADO_FINAL = DF_DERIVADO_FINAL.reset_index(drop=True)  

    ### Modelando dados do DF_BIG_TABLE
    DF_BIG_1 = DF_DERIVADO_FINAL.copy()
    DF_BIG_2 = DF_DISEL_FINAL.copy()
    DF_BIG_1 ['DATA'] = 'DADOS_DERIVADO' 
    DF_BIG_2 ['DATA'] = 'DADOS_DISEL'

    DF_BIG_FINAL = DF_BIG_1.append(DF_BIG_2, ignore_index=True)

    ### Salvando no Excel
    dir = './dados_desafio_airflow/new_tables'
    try:
        os.makedirs(dir)
        print('diretorio criado')
    except:
        print('diretorio já existe')

    # criar workbook
    wb = Workbook('./dados_desafio_airflow/new_tables/big_table.xlsx')
    wb.save('./dados_desafio_airflow/new_tables/big_table.xlsx')
    wb = load_workbook('./dados_desafio_airflow/new_tables/big_table.xlsx')


    # adicionar workbook
    sheet1 = wb.create_sheet('DF_DERIVADO_FINAL',0)
    sheet2 = wb.create_sheet('DF_DISEL_FINAL',1)
    sheet3 = wb.create_sheet('DF_BIG_FINAL',2)

    # Remove existing sheet
    ref = wb['Sheet']
    wb.remove(ref)

    # Seleciona sheet e exporta os dados
    active = wb['DF_DERIVADO_FINAL']
    for x in dataframe_to_rows(DF_DERIVADO_FINAL, index = False):
         active.append(x)

    active = wb['DF_DISEL_FINAL']
    for x in dataframe_to_rows(DF_DISEL_FINAL, index = False):
         active.append(x)
    active = wb['DF_BIG_FINAL']
    for x in dataframe_to_rows(DF_BIG_FINAL, index = False):
         active.append(x)

    # Salva a excel
    wb.save('./dados_desafio_airflow/new_tables/big_table.xlsx')
    return ('Concluido')

def enviando_email(ti):
        ### Email Teste
    DADOS_AT = ti.xcom_pull(task_ids='baixa_dados_externo')
    DADOS_AT = pd.DataFrame(DADOS_AT, columns =['at_data'])
    # Modulos para manipulação de email
    import smtplib
    import email, smtplib, ssl
    #import email
    import email, smtplib, ssl
    from email import encoders
    #import email.mime.application
    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    #import mimetypes
    import mimetypes

    ##### Senha do email Criptografado
    from cryptography.fernet import Fernet 
    with open('./Jupyter_Code/filekey.key', 'rb') as filekey:
        key = filekey.read()
        fernet = Fernet(key)
        token = fernet.decrypt(b'gAAAAABifI8x6ATV1HSDBBDCPAJ2yZTWCJryQvN-e7k1ZFvixWKhE9XIaaglT58RuHHR_Gcri4FW6FEgL6MSrx-WqHwqbveNmg==') 

    ### dados login servidor
    sender_email = "flowbot@tischool.net"
    #receiver_email = "weslleyfernand@gmail.com"
    receiver_email = "Pedro.Adamo@raizen.com"
    password = fernet.decrypt(b'gAAAAABifI8x6ATV1HSDBBDCPAJ2yZTWCJryQvN-e7k1ZFvixWKhE9XIaaglT58RuHHR_Gcri4FW6FEgL6MSrx-WqHwqbveNmg==')

    #######Dados do email se sofrer alteração
    subject = "Email com o Resultado do desafio em anexo"
    body = "Segue em anexo o big_table.xlsx"

    #######Dados se não sofrer alteração
    subject2 = "Relatorio do apn sem alteração"
    body2 = "Relatorio não sofreu ateração desde o dia " + DADOS_AT['at_data'][0]

    # Conexão dos dados de send/receiver
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    message["Bcc"] = receiver_email  # Recommended for mass emails

    # Corpo do email
    message.attach(MIMEText(body, "plain"))

    filename = "./dados_desafio_airflow/new_tables/big_table.xlsx"  # In same directory as script

    # Envio do arquivo em Binario
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode ASCII para envio do email   
    encoders.encode_base64(part)

    # Add header
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    # Add attachment
    message.attach(part)
    text = message.as_string()

    try:
        lastversion =pd.read_csv('lastversion.csv', sep = ';') 
        print(DADOS_AT)
        if lastversion['at_data'][0] != + DADOS_AT['at_data'][0] :
        
            # Log no servidor com context e envio do email
            context = ssl.create_default_context()
            with smtplib.SMTP("mail.tischool.net", 587) as server:
                server.login(sender_email, password.decode())
                server.sendmail(sender_email, receiver_email, text)
        else:
            # Conexão dos dados de send/receiver
            message2 = MIMEMultipart()
            message2["From"] = sender_email
            message2["To"] = receiver_email
            message2["Subject"] = subject2
            message2["Bcc"] = receiver_email  # Recommended for mass emails

            # Corpo do email
            message2.attach(MIMEText(body2, "plain"))
            #######Dados do email send/receiver

            text2 = message2.as_string()
            # Log no servidor com context e envio do email
            context = ssl.create_default_context()
            with smtplib.SMTP("mail.tischool.net", 587) as server:
                server.login(sender_email, password.decode())
                server.sendmail(sender_email, receiver_email, text2) 
            print('versão já enviada')
    except:

        # Log no servidor com context e envio do email
        context = ssl.create_default_context()
        with smtplib.SMTP("mail.tischool.net", 587) as server:
            server.login(sender_email, password.decode())
            server.sendmail(sender_email, receiver_email, text) 
    return ('Concluido')

def atualizando_lastversion(ti):
    DADOS_AT = ti.xcom_pull(task_ids='baixa_dados_externo')
    DADOS_AT = pd.DataFrame(DADOS_AT, columns =['at_data'])
    lastversion = DADOS_AT
    lastversion.to_csv('./dados_desafio_airflow/lastversion.csv', sep = ';')
    return ('Concluido')

default_args = {
    'owner': 'XYZ',
    'start_date': datetime(2022, 5, 11),
    'schedule_interval': '@daily',
}

with DAG('Desafio_airflow', catchup=False, default_args=default_args) as dag:


    baixa_dados_externo = PythonOperator(
        task_id = 'baixa_dados_externo',
        python_callable = baixa_dados_externo
    )


    confirmando_data_atualizacao = PythonOperator(
        task_id = 'confirmando_data_atualizacao',
        python_callable = confirmando_data_atualizacao
    )

    tratando_dados_save_bigdate = PythonOperator(
        task_id = 'tratando_dados_save_bigdate',
        python_callable = tratando_dados_save_bigdate
    )
    enviando_email = PythonOperator(
        task_id = 'enviando_email',
        python_callable = enviando_email
    )


    atualizando_lastversion = PythonOperator(
        task_id = 'atualizando_lastversion',
        python_callable = atualizando_lastversion
    )        

baixa_dados_externo >> confirmando_data_atualizacao >> tratando_dados_save_bigdate >> enviando_email >> atualizando_lastversion